import pandas as pd
from connection import *
from openpyxl import load_workbook
import tkinter
from datetime import datetime, timedelta

base_url = "SEU_IP:PORTA_UNIFI_AQUI"


def login(session, username, password):
    login_url = f"{base_url}/api/login"
    data = {"username": username, "password": password}
    response = session.post(login_url, json=data, verify=False)

    if response.status_code == 200:
        return True
    else:
        return response


def create_voucher(session, dias, quant, dispo, setor, nome):
    notas = ' - '.join([nome, setor, dispo])
    create_url = f"{base_url}/api/s/default/cmd/hotspot"
    payload = {
        "cmd": "create-voucher",
        "expire": dias * 1440,
        "n": quant,
        "note": notas
    }
    headers = {"Content-Type": "application/json"}
    response = session.post(create_url, json=payload, headers=headers, verify=False)
    if response.status_code == 200:
        created_time = response.json()
        created_time = created_time['data'][0]['create_time']
        return response, created_time
    else:
        return response, response.text

def detalhes_voucher(session, create_time):
    vouchers = []
    detalhes_url = f"{base_url}/api/s/default/stat/voucher"
    payload = {
        "create_time": create_time
    }
    response = session.get(detalhes_url, json=payload, verify=False)
    data = response.json()

    for n in data['data']:
        vouchers.append(n['code'])

    return vouchers


def revoke_voucher(session, id):
    revoke_url = f"{base_url}/api/s/default/cmd/hotspot"
    payload = {
        '_id': id,
        'cmd': 'delete-voucher'
    }
    response = session.post(revoke_url, json=payload, verify=False)
    if response.status_code == 200:
        tkinter.messagebox.showinfo("Resultado", f"Voucher apagado")
    else:
        tkinter.messagebox.showinfo("Resultado", f"Erro ao apagar voucher")


def extend_voucher(session, id, dias):
    n = 0
    extend_url = f"{base_url}/api/s/default/cmd/hotspot/guests"

    payload = {
        '_id': id,
        'cmd': 'extend'
    }

    while n < dias:
        response = session.post(extend_url, json=payload, verify=False)
        n += 1
        
    if response.status_code == 200:
        tkinter.messagebox.showinfo("Resultado", f"Voucher estendido em {dias} dias")
    else:
        tkinter.messagebox.showinfo("Resultado", f"Erro ao estender voucher: {response.text}")


def list_guests(session, within=1):
    guest_url = f"{base_url}/api/s/default/stat/guest"
    payload = {
        'within': within
    }

    response = session.post(guest_url, json=payload, verify=False)
    data = response.json()

    return data


def criar_planilhar(data):
    guests = []

    for n in data['data']:
        guests.append([n.get('name', 'N/A'),
                       n.get('voucher_code', 'N/A'),
                       n.get('end', 'N/A'),
                       n.get('_id', 'N/A')])

    df = pd.DataFrame(guests, columns=['Nome', 'Voucher Code', 'Dias Restantes', 'ID'])
    df.to_excel('guests.xlsx', index=False)

    book = load_workbook('guests.xlsx')
    sheet = book.active

    for row in range(2, len(guests) + 2):
        cell = f'E{row}'
        formula = f'=ARREDONDAR.PARA.CIMA(C{row}/86400+25569, 0)'
        sheet[cell] = formula

    book.save('guests_formulas.xlsx')


def salvar_planilha(vouchers, nome, setor, dispo, dias):
    try:
        wb = load_workbook(r'T:\TI\Informatica\Controles\Controle de Vouchers MSP.xlsx')
        ws = wb['FUNCIONÁRIOS']

        hoje = datetime.now()
        # Adicionar 75 dias à data de hoje
        data_futura = hoje + timedelta(days=int(dias))
        data_futura = data_futura.strftime("%d/%m/%Y")

        for cell in ws['B']:
            if cell.value is None:
                row = cell.row
                for numero in vouchers:
                    ws.cell(row=int(row), column=2, value=numero)
                    ws.cell(row=int(row), column=3, value=nome)
                    ws.cell(row=int(row), column=4, value=setor)
                    ws.cell(row=int(row), column=5, value=dispo)
                    ws.cell(row=int(row), column=6, value=data_futura)
                    row += 1
                break

        wb.save(r'T:\TI\Informatica\Controles\Controle de Vouchers MSP.xlsx')

    except PermissionError as e:
        resposta = tkinter.messagebox.askyesno("Não foi possível colocar na planilha",
                                               f"{str(e)}\n\nNão foi possível colocar os dados do voucher na planilha. Possíveis causas: \n 1 - A planilha já está aberta \n 2 - Esse aplicativo não tem permissão para alterar a planilha\n\nTentar novamente?")
        return resposta


def criar_txtfile(vouchers):
    with open('numeros_voucher.txt', 'w') as arquivo:
        for numero in vouchers:
            arquivo.write(numero + '\n')

def inserir_list_guest(session, cursor, conn):
    data = list_guests(session)
    data = data['data']

    campos_selecionados = []
    for item in data:
        campos_selecionados.append({
            'nome': item.get('name'),
            'id': item.get('_id'),
            'code': item.get('voucher_code')
        })

    for item in campos_selecionados:
        cursor.execute('''
        INSERT OR REPLACE INTO users (id, nome, code) VALUES (?, ?, ?)
        ''', (item['id'], item['nome'], item['code']))

    conn.commit()

    conn.close()