import tkinter.messagebox
import requests
import customtkinter as ctk
from datetime import datetime, timedelta
from openpyxl import load_workbook

class App(ctk.CTk):

    def __init__(self):
        super().__init__()
        #URL BASE
        self.base_url = "https://SEU_IP_UNIFI_AQUI" #ONDE ESTÁ HOSPEDADO O UNIFI
        self.session = None
        self.vouchers = None
        self.login = ''
        self.senha = ''
        self.configuracao_layout()
        self.janela_inicial()

    def configuracao_layout(self):
        self.title('Gerador de Voucher')
        self.geometry('500x300')
        self.resizable(False, False)

    def janela_inicial(self):
        # primeiro frame
        frame = ctk.CTkFrame(self, width=500, height=50, corner_radius=0, bg_color='teal', fg_color='#30A6D9')
        frame.place(x=0, y=10)
        ctk.CTkLabel(frame, text='Gerador de Voucher', font=('Century Gothic bold', 20),
                     text_color='#fff').place(x=165, y=12)
        # frame_login
        self.frame_login = ctk.CTkFrame(self, width=500, height=200, corner_radius=0, bg_color='transparent',
                                        fg_color='transparent')
        self.frame_login.place(x=0, y=60)
        ctk.CTkLabel(self.frame_login, text='Login', font=('Century Gothic bold', 16)).place(x=230, y=5)
        login_entry = ctk.CTkEntry(self.frame_login)
        login_entry.place(x=180, y=35)
        ctk.CTkLabel(self.frame_login, text='Senha', font=('Century Gothic bold', 16)).place(x=230, y=75)
        senha_entry = ctk.CTkEntry(self.frame_login, show='*')
        senha_entry.place(x=180, y=105)
        validacao_login = ctk.CTkLabel(self.frame_login, text='', font=('Century Gothic bold', 16), text_color='red')
        validacao_login.place(x=160, y=135)

        ctk.CTkButton(self.frame_login, text='LOGIN', command=lambda: (login := login_entry.get(),
                                                                        senha := senha_entry.get(),
                                                                        self.logar(login, senha)),
                      fg_color='#30A6D9', hover_color='#3071D9', font=('Century Gothic bold', 14)).place(x=180, y=170)

    def janela_gerador(self):
        text_var = ctk.BooleanVar()
        plan_var = ctk.BooleanVar()

        self.frame_gerar = ctk.CTkFrame(self, width=500, height=200, corner_radius=0, bg_color='transparent',
                                        fg_color='transparent')
        self.frame_gerar.place(x=0, y=10)
        ctk.CTkLabel(self.frame_gerar, text='Dias', font=('Century Gothic bold', 16)).place(x=85, y=5)
        dias_entry = ctk.CTkEntry(self.frame_gerar, width=100)
        dias_entry.place(x=55, y=35)
        ctk.CTkLabel(self.frame_gerar, text='Quantidade', font=('Century Gothic bold', 16)).place(x=210, y=5)
        quant_entry = ctk.CTkEntry(self.frame_gerar, width=100)
        quant_entry.place(x=200, y=35)
        ctk.CTkLabel(self.frame_gerar, text='Setor', font=('Century Gothic bold', 16)).place(x=375, y=5)
        setor_entry = ctk.CTkEntry(self.frame_gerar, width=100)
        setor_entry.place(x=345, y=35)
        ctk.CTkLabel(self.frame_gerar, text='Nome', font=('Century Gothic bold', 16)).place(x=345, y=75)
        nome_entry = ctk.CTkEntry(self.frame_gerar, width=150)
        nome_entry.place(x=295, y=105)
        ctk.CTkLabel(self.frame_gerar, text='Dispositivo', font=('Century Gothic bold', 16)).place(x=120, y=75)
        dispo_entry = ctk.CTkEntry(self.frame_gerar, width=210)
        dispo_entry.place(x=55, y=105)
        self.validacao_gerador = ctk.CTkLabel(self.frame_gerar, text='', font=('Century Gothic bold', 16),
                                              text_color='red')
        self.validacao_gerador.place(x=100, y=135)
        self.criar_plan = ctk.CTkCheckBox(self.frame_gerar, text='Salvar na planilha', onvalue=True, offvalue=False,
                                          variable=plan_var, border_width=1)
        self.criar_plan.place(x=55, y=168)
        self.criar_txt = ctk.CTkCheckBox(self.frame_gerar, text='Criar TXT', onvalue=True, offvalue=False,
                                         variable=text_var, border_width=1)
        self.criar_txt.place(x=190, y=168)
        botao_gerar = ctk.CTkButton(self.frame_gerar, text='GERAR', command=lambda: (dias := dias_entry.get(),
                                                                                     quant := quant_entry.get(),
                                                                                     dispo := dispo_entry.get(),
                                                                                     setor := setor_entry.get(),
                                                                                     nome := nome_entry.get(),
                                                                                     self.create_voucher(dias,
                                                                                                quant, dispo, setor,
                                                                                                nome, text_var,
                                                                                                plan_var) if dias != '' and quant != '' and dispo != '' and nome != '' else self.validacao_gerador.configure(
                                                                                         text='Todos campos são obrigatórios')),
                                    fg_color='#30A6D9',
                                    hover_color='#3071D9', font=('Century Gothic bold', 14), width=150)
        botao_gerar.place(x=300, y=165)

    def logar(self, username, password):
        self.login = username
        self.senha = password
        # Endpoint de login
        login_url = f"{self.base_url}/api/login"

        # Dados para o corpo da solicitação POST
        data = {"username": self.login, "password": self.senha}

        # Criando uma sessão
        self.session = requests.Session()

        # Fazendo a solicitação POST para fazer login
        response = self.session.post(login_url, json=data, verify=False)

        # Verificando se o login foi bem-sucedido
        if response.status_code == 200:
            self.frame_login.destroy()
            self.janela_gerador()
        elif response.status_code == 400:
            tkinter.messagebox.showinfo('Erro de credencial', 'Login ou senha incorretos')
        else:
            tkinter.messagebox.showinfo('Erro', 'Erro ao fazer login: ', response.text)

    def create_voucher(self, dias, quant, dispo, setor, nome, text_var, plan_var):
        notas = ' - '.join([nome, setor, dispo])
        create_url = f"{self.base_url}/api/s/default/cmd/hotspot"
        payload = {
                "cmd": "create-voucher",
                "expire": int(dias) * 1440,
                "n": quant,
                "note": notas
            }
        print(dias)
        headers = {"Content-Type": "application/json"}
        response = self.session.post(create_url, json=payload, headers=headers, verify=False)
        if response.status_code == 200:
            created_time = response.json()
            created_time = created_time['data'][0]['create_time']
            hoje = datetime.now()
            # Adicionar 75 dias à data de hoje
            data_futura = hoje + timedelta(days=int(dias))
            data_futura = data_futura.strftime("%d/%m/%Y")
            self.detalhes_voucher(created_time, data_futura, nome, setor, dispo, plan_var, text_var)

        else:
            tkinter.messagebox.showinfo('Erro', f'Erro: {response.text}')

    def detalhes_voucher(self, create_time, data_futura, nome, setor, dispo, plan_var, text_var):
        vouchers = []
        detalhes_url = f"{self.base_url}/api/s/default/stat/voucher"
        payload = {
            "create_time": create_time
        }
        response = self.session.get(detalhes_url, json=payload, verify=False)
        data = response.json()
        for n in data['data']:
            vouchers.append(n['code'])

        self.vouchers = vouchers



        #Tela com vouchers
        def copy_text(event):
            text = event.widget.cget("text")
            event.widget.focus_set()
            event.widget.clipboard_clear()
            event.widget.clipboard_append(text)

        def criar_label(frame, texto):
            label = ctk.CTkLabel(frame, text=texto)
            label.pack()
            label.configure(cursor="hand2")
            label.bind("<Button-1>", copy_text)

        def fechar_vouchers():
            self.frame_vouchers.destroy()
            self.logar(self.login, self.senha)

        def apresentar_vouchers():
            self.frame_vouchers = ctk.CTkScrollableFrame(self, width=500, height=200, corner_radius=0,
                                                         bg_color='transparent',
                                                         fg_color='transparent')
            self.frame_vouchers.place(x=0, y=10)
            for voucher in self.vouchers:
                criar_label(self.frame_vouchers, voucher)

            ctk.CTkButton(self.frame_vouchers, text='FECHAR', command=fechar_vouchers,
                          fg_color='#30A6D9', hover_color='#3071D9', font=('Century Gothic bold', 14)).pack()
        apresentar_vouchers()

        def salvar_planilha():
            try:
                wb = load_workbook(r'CAMINHO PARA ABRIR CARREGAR A PLANILHA')
                ws = wb['FUNCIONÁRIOS']

                for cell in ws['B']:
                    if cell.value is None:
                        row = cell.row
                        for numero in vouchers:
                            ws.cell(row=int(row), column=2, value=numero)
                            ws.cell(row=int(row), column=3, value=nome)
                            ws.cell(row=int(row), column=4, value=setor)
                            ws.cell(row=int(row), column=5, value=dispo)
                            ws.cell(row=int(row), column=6, value=data_futura)
                            row += 1
                        break

                wb.save(r'CAMINHO PARA SALVAR')

            except PermissionError as e:
                resposta = tkinter.messagebox.askyesno("Não foi possível colocar na planilha",
                                                       f"{str(e)}\n\nNão foi possível colocar os dados do voucher na planilha. Possíveis causas: \n 1 - A planilha já está aberta \n 2 - Esse aplicativo não tem permissão para alterar a planilha\n\nTentar novamente?")

                if resposta:
                    salvar_planilha()

        if plan_var.get():
            salvar_planilha()

        if text_var.get():
            with open('numeros_voucher.txt', 'w') as arquivo:
                for numero in vouchers:
                    arquivo.write(numero + '\n')

if __name__ == '__main__':
    app = App()
    app.mainloop()